import streamlit as st
# from io import BytesIO
import pandas as pd
# import datetime
# from datetime import date
# import numpy as np
# import xlsxwriter
import warnings
warnings.simplefilter("ignore")
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

st.set_page_config(layout="wide", page_icon=":clipboard:", page_title="Modules Report")

# This is the title to the code
st.title('Modules Report')
st.write('This report details the number of modules that were sat in a given month.')
st.write('It also has a dashboard that gives you a count as to the number of modules sat and the average score for the modules sat.')
st.warning('It does not record fails.')
# ... (the rest of your code remains the same) ...
# This is the uploader for the excel file.
uploaded_file1 = st.file_uploader("Upload Modules Report", key='1')
if uploaded_file1 is not None:
    df1 = pd.read_excel(uploaded_file1)

if st.button('Create Dashboard', key='2'):
    st.success('Dashboard Completed')
    workbook = openpyxl.Workbook()

    # Create a new worksheet for the original data
    original_data_sheet = workbook.active
    original_data_sheet.title = 'Original Data'

    # Write the dataframe to the 'Original Data' worksheet
    for row in dataframe_to_rows(df1, index=False, header=True):
        original_data_sheet.append(row)

    # Apply formatting to the 'Original Data' worksheet
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='left', vertical='center', indent=1)
    cell_alignment = Alignment(horizontal='left', vertical='center', indent=1)

    for row in original_data_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=df1.shape[1]):
        for cell in row:
            cell.font = header_font
            cell.alignment = header_alignment

    for row in original_data_sheet.iter_rows(min_row=2, max_row=original_data_sheet.max_row, min_col=1, max_col=df1.shape[1]):
        for cell in row:
            cell.alignment = cell_alignment

    # Set the height of each row in the 'Dashboard' worksheet
    #original_data_sheet.row_dimensions[1].height = 25

    for row in range(1, original_data_sheet.max_row + 1):
        original_data_sheet.row_dimensions[row].height = 25

    original_data_sheet.column_dimensions['A'].width = 10
    original_data_sheet.column_dimensions['B'].width = 10
    original_data_sheet.column_dimensions['C'].width = 10
    original_data_sheet.column_dimensions['D'].width = 10
    original_data_sheet.column_dimensions['E'].width = 20
    original_data_sheet.column_dimensions['F'].width = 15
    original_data_sheet.column_dimensions['G'].width = 25

    # Create a new worksheet named 'Dashboard'
    dashboard_sheet = workbook.create_sheet('Dashboard')
    # Format the 'Dashboard' sheet
    dashboard_sheet.row_dimensions[1].height = 25
    dashboard_sheet.row_dimensions[2].height = 25
    dashboard_sheet.column_dimensions['A'].width = 20
    dashboard_sheet.column_dimensions['B'].width = 20

    # Count the number of modules in column A of the 'Original Data' worksheet
    module_count = df1['User ID'].count()

    # Calculate the average score in column B of the 'Original Data' worksheet
    average_score = df1['Grade'].mean()

    # Write the module count and average score to the 'Dashboard' worksheet
    dashboard_sheet['A1'] = 'Module Count'
    dashboard_sheet['A2'] = module_count
    dashboard_sheet['B1'] = 'Average Score'
    dashboard_sheet['B2'] = average_score

    # Apply formatting to the 'Dashboard' worksheet
    for row in dashboard_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
        for cell in row:
            cell.font = header_font
            cell.alignment = header_alignment

    for row in dashboard_sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = cell_alignment

    # Save the modified workbook
    workbook.save('Monthly_Modules_Report.xlsx')

    st.download_button(
        label="Download data as XLSX",
        data=open('Monthly_Modules_Report.xlsx', 'rb'),
        file_name='Monthly_Modules_Report.xlsx'
    )   
    


    


    
